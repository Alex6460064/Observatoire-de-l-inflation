"""Collecte du salaire mensuel de base (SMB) -- Dares, enquete trimestrielle
Acemo.

Source : docs/SOURCES.md, section "DARES -- Salaire mensuel de base (SMB),
comparaison salaires/prix". Feuille `Sal. mens. ensemble` du classeur publie
sur <https://dares.travail-emploi.gouv.fr/donnees/les-indices-de-salaire-de-base>,
ligne `ENS` ("Ensemble des secteurs non agricoles").

Le domaine `dares.travail-emploi.gouv.fr` est protege par une verification
anti-bot Cegedim qui bloque `requests`/`curl`, meme avec un `User-Agent` de
navigateur -- verifie le 23/08/2026 (docs/adr/0022). Ce module ne fait donc
aucun appel reseau : il lit un fichier deja telecharge a la main et archive
dans `data/raw/` (patron de l'ADR 0004, meme traitement que Familles Rurales
et ARCEP). Le nom du fichier change a chaque publication trimestrielle --
toujours repartir de la page ci-dessus, jamais d'une URL de fichier codee en
dur (les mentions legales Dares le disent explicitement).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

FEUILLE_ENSEMBLE = "Sal. mens. ensemble"
CODE_POSTE_ENSEMBLE = "ENS"

MOIS_VERS_NUMERO = {"mars": "03", "juin": "06", "sept": "09", "dec": "12"}


def chemin_dernier_fichier_salaire_smb(raw_dir: Path = Path("data/raw")) -> Path:
    """Le plus recent fichier archive localement pour cette source.

    Args:
        raw_dir: dossier de recherche, motif `dares_salaire_smb_AAAA-MM-JJ.xlsx`.

    Returns:
        Le chemin du fichier le plus recent (tri lexicographique = tri par
        date, format ISO dans le nom).

    Raises:
        FileNotFoundError: aucun fichier archive -- telecharger a la main
            depuis
            https://dares.travail-emploi.gouv.fr/donnees/les-indices-de-salaire-de-base
            (ADR 0022 : le blocage anti-bot Cegedim empeche l'automatisation).
    """
    fichiers = sorted(raw_dir.glob("dares_salaire_smb_*.xlsx"))
    if not fichiers:
        raise FileNotFoundError(
            f"Aucun fichier 'dares_salaire_smb_*.xlsx' dans {raw_dir}. "
            "Telecharger a la main depuis "
            "https://dares.travail-emploi.gouv.fr/donnees/"
            "les-indices-de-salaire-de-base "
            "et l'archiver sous ce nom (ADR 0022)."
        )
    return fichiers[-1]


def lire_salaire_smb(chemin_xlsx: Path) -> pd.DataFrame:
    """Parse le classeur Dares : feuille `Sal. mens. ensemble`, ligne `ENS`.

    Args:
        chemin_xlsx: fichier deja telecharge a la main (voir
            `chemin_dernier_fichier_salaire_smb`).

    Returns:
        Table longue `source, poste, periode, valeur` -- `source="dares"`,
        `poste="ENS"`, `periode` au format `AAAA-MM` (mois de fin de
        trimestre), `valeur` l'indice SMB tel que publie, base 100 en juin
        2017 (pas encore rebase sur la reference du projet). Les trimestres
        marques `n.d.` dans le fichier (non encore publies) sont absents de
        la table plutot que mis a `NaN` -- `traitement.dares.completer_mensuel`
        les reconstitue par interpolation (ADR 0015).

    Raises:
        ValueError: feuille, ligne `ENS`, ou en-tete annee/mois absents ou
            de forme inattendue -- le schema Dares a pu changer depuis la
            verification du 23/08/2026, voir docs/SOURCES.md avant de
            corriger silencieusement.
    """
    classeur = openpyxl.load_workbook(chemin_xlsx, data_only=True, read_only=True)
    if FEUILLE_ENSEMBLE not in classeur.sheetnames:
        raise ValueError(
            f"Feuille '{FEUILLE_ENSEMBLE}' absente de {chemin_xlsx} "
            f"(feuilles presentes : {classeur.sheetnames}). "
            "Le classeur Dares a peut-etre change de structure."
        )
    lignes_feuille = list(classeur[FEUILLE_ENSEMBLE].iter_rows(values_only=True))

    idx_mois = _trouver_ligne_mois(lignes_feuille, chemin_xlsx)
    ligne_annee = lignes_feuille[idx_mois + 1]
    ligne_ens = _trouver_ligne_ens(lignes_feuille, idx_mois + 2, chemin_xlsx)
    ligne_mois = lignes_feuille[idx_mois]

    resultats = []
    for col in range(2, len(ligne_annee)):
        annee = ligne_annee[col]
        if not isinstance(annee, int):
            break  # fin des colonnes trimestrielles (section "Variations sur")
        mois = ligne_mois[col]
        if mois not in MOIS_VERS_NUMERO:
            raise ValueError(
                f"Mois inattendu en colonne {col} de {chemin_xlsx} : {mois!r}. "
                "Le schema Dares a peut-etre change."
            )
        valeur = ligne_ens[col]
        if valeur is None or valeur == "n.d.":
            continue  # trimestre non encore publie
        resultats.append(
            (
                "dares",
                CODE_POSTE_ENSEMBLE,
                f"{annee}-{MOIS_VERS_NUMERO[mois]}",
                float(valeur),
            )
        )

    if not resultats:
        raise ValueError(
            f"Aucune valeur exploitable dans {chemin_xlsx}, feuille "
            f"'{FEUILLE_ENSEMBLE}'. Le schema Dares a peut-etre change."
        )

    return pd.DataFrame(resultats, columns=["source", "poste", "periode", "valeur"])


def _trouver_ligne_mois(lignes: list[tuple], chemin_xlsx: Path) -> int:
    for i, ligne in enumerate(lignes):
        valeurs = set(ligne[2:6]) if len(ligne) >= 6 else set()
        if valeurs & set(MOIS_VERS_NUMERO):
            return i
    raise ValueError(
        f"Ligne d'en-tete des mois introuvable dans {chemin_xlsx}, feuille "
        f"'{FEUILLE_ENSEMBLE}'. Le schema Dares a peut-etre change."
    )


def _trouver_ligne_ens(lignes: list[tuple], depart: int, chemin_xlsx: Path) -> tuple:
    for ligne in lignes[depart:]:
        if ligne and ligne[0] == CODE_POSTE_ENSEMBLE:
            return ligne
    raise ValueError(
        f"Ligne '{CODE_POSTE_ENSEMBLE}' introuvable dans {chemin_xlsx}, feuille "
        f"'{FEUILLE_ENSEMBLE}'. Le schema Dares a peut-etre change."
    )
